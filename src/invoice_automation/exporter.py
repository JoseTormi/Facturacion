from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from invoice_automation.models import Invoice


def export_invoices_to_excel(invoices: list[Invoice], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = [invoice.to_excel_row() for invoice in invoices]
    dataframe = pd.DataFrame(rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="Facturas")
        worksheet = writer.sheets["Facturas"]
        format_header(worksheet)
        autosize_columns(worksheet)

    return output_path


def format_header(worksheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)
